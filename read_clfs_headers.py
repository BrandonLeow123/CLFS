r"""
read_clfs_headers.py

Simple standalone script to read an Excel file (default: CLFS_sample_input.xlsx)
and print the sheet names, column headers (with indices), and a sanitized headers
dictionary mapping python-friendly keys -> { 'index', 'header' }.

Usage (PowerShell):
  python read_clfs_headers.py --path "C:\Users\brand\Desktop\CLFS\CLFS\CLFS_sample_input (3).xlsx"

This script intentionally keeps behavior minimal (no file writes) — tell me if you'd
like JSON/CSV export or automatic mapping to a DataFrame.
"""

from __future__ import annotations
import argparse
import os
import re
from typing import Dict, Any, Tuple, Optional

import pandas as pd
import zipfile
import xml.etree.ElementTree as ET
from clfs_headers_mapping import HEADERS_MAPPING
import re
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except Exception:
    load_workbook = None
    PatternFill = None


def make_safe_key(s: str) -> str:
    """Create a sanitized, python-friendly key from a header string.

    - lowercases
    - replaces whitespace with underscore
    - removes non-alphanumeric/underscore
    - collapses multiple underscores
    """
    k = (s or "").lower()
    k = re.sub(r"\s+", "_", k)
    k = re.sub(r"[^0-9a-z_]+", "", k)
    k = re.sub(r'_+', '_', k)
    k = k.strip('_')
    return k or 'col'


def read_headers_from_excel(path: str, sheet_name=0, engine: str = 'openpyxl') -> Tuple[pd.DataFrame | None, Dict[str, Dict[str, Any]], Optional[int]]:
    """Open an Excel file and read headers from the given sheet.

    Returns (df, headers_dict, header_row). df may be None if reading failed.
    headers_dict maps sanitized key -> {'index': idx, 'header': original_header}
    """
    path = os.path.expanduser(path)
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    # list sheets
    try:
        xls = pd.ExcelFile(path, engine=engine)
        print('Sheets found:', xls.sheet_names)
    except Exception as e:
        print('Could not read Excel file metadata:', e)
        xls = None

    try:
        df = pd.read_excel(path, sheet_name=sheet_name, header=0, engine=engine)
        cols = list(df.columns)
        print(f"\nRead sheet {sheet_name!r} — {len(cols)} columns")
        for i, c in enumerate(cols, 1):
            name = str(c).strip()
            display = '<EMPTY>' if name == '' or name.lower() == 'nan' else name
            print(f"{i:03d}: {display}")

        # build sanitized keys with uniqueness enforcement
        headers_dict: Dict[str, Dict[str, Any]] = {}
        counts: Dict[str, int] = {}
        for idx, h in enumerate(cols):
            base = make_safe_key(str(h))
            counts[base] = counts.get(base, 0) + 1
            key = base if counts[base] == 1 else f"{base}_{counts[base]}"
            headers_dict[key] = {'index': idx, 'header': h}

        # header_row unknown when using pandas.read_excel (header=0)
        return df, headers_dict, None

    except Exception as e:
        print('Failed to read with header=0:', e)
        print('Falling back to XML-based header detection and read (more robust).')
        # Attempt to detect header row via XML parsing (works even when openpyxl/pandas struggles)
        try:
            header_row, header_values = detect_header_row_from_xml(path)
            print(f"Detected header row: {header_row} (Excel numbering)")
            df_xml = read_sheet_from_xml(path, header_row)
            # build sanitized mapping from detected header_values
            headers_dict = {}
            counts: Dict[str, int] = {}
            for idx, h in enumerate(header_values):
                base = make_safe_key(str(h))
                counts[base] = counts.get(base, 0) + 1
                key = base if counts[base] == 1 else f"{base}_{counts[base]}"
                headers_dict[key] = {'index': idx, 'header': h}
            return df_xml, headers_dict, header_row
        except Exception as exml:
            print('XML fallback also failed:', exml)
            print('Attempting to read top rows with header=None for inspection...')
            try:
                df2 = pd.read_excel(path, sheet_name=sheet_name, header=None, nrows=10, engine=engine)
                print('\nTop rows (header=None):')
                print(df2.to_string(index=False, header=False))
            except Exception as e2:
                print('Also failed reading with header=None:', e2)
        return None, {}, None


def _norm_simple(s: str) -> str:
    if s is None:
        return ''
    s = str(s).strip().lower()
    s = re.sub(r'[^0-9a-z ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def detect_header_row_from_xml(path: str) -> Tuple[int, list]:
    """Return the Excel header row number (1-based) and the header values list by inspecting
    xl/sharedStrings.xml and xl/worksheets/sheet1.xml. Chooses the row with the most
    matches against HEADERS_MAPPING."""
    with zipfile.ZipFile(path, 'r') as z:
        ss = z.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
        sst = ET.fromstring(ss)
        ns = {'m': 'http://purl.oclc.org/ooxml/spreadsheetml/main'}
        shared = []
        for si in sst.findall('m:si', ns):
            t = si.find('m:t', ns)
            if t is not None:
                shared.append(t.text)
                continue
            parts = []
            for r in si.findall('m:r', ns):
                rt = r.find('m:t', ns)
                if rt is not None:
                    parts.append(rt.text)
            shared.append(''.join(p for p in parts if p))

        sh = z.read('xl/worksheets/sheet1.xml').decode('utf-8', errors='replace')
    root = ET.fromstring(sh)
    ns_s = {'s': 'http://purl.oclc.org/ooxml/spreadsheetml/main'}

    mapping_headers = [entry['header'] for entry in HEADERS_MAPPING.values()]
    mapping_norm = { _norm_simple(h): h for h in mapping_headers }

    best = (0.0, 0, [])  # (score, rownum, header_list)
    for row in root.findall('s:sheetData/s:row', ns_s):
        rnum = int(row.attrib.get('r', '0'))
        vals = []
        for c in row.findall('s:c', ns_s):
            t = c.attrib.get('t')
            v = c.find('s:v', ns_s)
            if v is None:
                vals.append(None)
            else:
                if t == 's':
                    idx = int(v.text)
                    vals.append(shared[idx] if idx < len(shared) else None)
                else:
                    vals.append(v.text)
        norm_vals = set(_norm_simple(v) for v in vals if v is not None)
        matches = sum(1 for k in mapping_norm.keys() if k and k in norm_vals)
        score = matches / max(1, len(vals))
        if matches > best[1] or (matches == best[1] and score > best[0]):
            best = (score, matches, (rnum, vals))

    if best[2]:
        rnum, header_vals = best[2]
        return rnum, [v if v is not None else '' for v in header_vals]
    raise ValueError('Could not detect header row from XML')


def read_sheet_from_xml(path: str, header_row: int) -> pd.DataFrame:
    """Read the worksheet from the .xlsx ZIP and return a pandas.DataFrame using the
    provided header_row (1-based). Only supports simple cell types and sharedStrings."""
    with zipfile.ZipFile(path, 'r') as z:
        ss = z.read('xl/sharedStrings.xml').decode('utf-8', errors='replace')
        sst = ET.fromstring(ss)
        ns = {'m': 'http://purl.oclc.org/ooxml/spreadsheetml/main'}
        shared = []
        for si in sst.findall('m:si', ns):
            t = si.find('m:t', ns)
            if t is not None:
                shared.append(t.text)
                continue
            parts = []
            for r in si.findall('m:r', ns):
                rt = r.find('m:t', ns)
                if rt is not None:
                    parts.append(rt.text)
            shared.append(''.join(p for p in parts if p))

        sh = z.read('xl/worksheets/sheet1.xml').decode('utf-8', errors='replace')
    root = ET.fromstring(sh)
    ns_s = {'s': 'http://purl.oclc.org/ooxml/spreadsheetml/main'}

    # build rows as dicts keyed by 1-based column index
    rows = {}
    maxcol = 0
    for row in root.findall('s:sheetData/s:row', ns_s):
        rnum = int(row.attrib.get('r', '0'))
        cells = {}
        for c in row.findall('s:c', ns_s):
            coord = c.attrib.get('r')  # e.g., A6
            m = re.match(r'([A-Z]+)(\d+)', coord)
            if not m:
                continue
            colletters = m.group(1)
            colidx = 0
            for ch in colletters:
                colidx = colidx*26 + (ord(ch) - 64)
            t = c.attrib.get('t')
            v = c.find('s:v', ns_s)
            val = None
            if v is not None:
                if t == 's':
                    idx = int(v.text)
                    val = shared[idx] if idx < len(shared) else None
                else:
                    val = v.text
            cells[colidx] = val
            if colidx > maxcol:
                maxcol = colidx
        rows[rnum] = cells

    # header values
    header_cells = rows.get(header_row)
    if not header_cells:
        raise ValueError(f'Header row {header_row} not found in sheet XML')
    headers = [header_cells.get(i, '') for i in range(1, maxcol+1)]

    data = []
    for r in sorted(k for k in rows.keys() if k > header_row):
        rowvals = [rows[r].get(i) for i in range(1, maxcol+1)]
        data.append(rowvals)

    df = pd.DataFrame(data, columns=headers)
    return df


def main():
    parser = argparse.ArgumentParser(description='Read Excel headers for CLFS sample file')
    parser.add_argument('--path', '-p', help='Path to Excel file', default='CLFS_sample_input (3).xlsx')
    parser.add_argument('--sheet', '-s', help='Sheet name or index (default 0)', default=0)
    parser.add_argument('--highlight-output', '-o', help='If provided, write a copy of the workbook with flagged rows highlighted')
    args = parser.parse_args()

    path = args.path
    sheet = args.sheet
    # coerce sheet to int when possible
    try:
        sheet_val = int(sheet)
    except Exception:
        sheet_val = sheet

    try:
        df, headers, header_row = read_headers_from_excel(path, sheet_name=sheet_val)
        if headers:
            print('\nSanitized header keys (example, first 50):')
            keys = list(headers.keys())
            for i, k in enumerate(keys[:50], 1):
                print(f"{i:03d}. {k} -> {headers[k]['header']}")
            # print a small preview using the detected headers
            try:
                print('\nPreview (first 5 rows):')
                pd.set_option('display.max_colwidth', 120)
                if df is not None and not df.empty:
                    print(df.head(5).to_string(index=False))
                else:
                    print('No data rows found after header row (empty dataframe).')
                # show any columns present in df that are not in the mapping (unmapped)
                df_cols = [str(c) for c in (df.columns if df is not None else [])]
                mapped_headers = [v['header'] for v in headers.values()]
                unmapped = [c for c in df_cols if c not in mapped_headers]
                if unmapped:
                    print('\nUnmapped columns (present in sheet, not in HEADERS_MAPPING):')
                    for c in unmapped[:50]:
                        print(' -', c)
            except Exception as pe:
                print('Failed to print preview:', pe)

            # validation rule: if internship == Yes then type_of_employment must be 'Fixed-Term Contract employee'
            try:
                # Use column indices rather than labels to avoid ambiguous column label selection
                key_intern = 'was_your_main_job_last_week_a_paid_internship_traineeship_or_apprenticeship'
                key_type = 'type_of_employment'
                if key_intern in headers and key_type in headers and df is not None:
                    col_intern_idx = headers[key_intern]['index']
                    col_type_idx = headers[key_type]['index']
                    # grab exact Series by integer location to avoid duplicate-label ambiguity
                    s_intern = df.iloc[:, col_intern_idx].astype(str).str.strip().str.lower()
                    s_type = df.iloc[:, col_type_idx].astype(str).str.strip().str.lower()
                    intern_yes = s_intern.isin(['yes', 'y', 'true'])
                    is_fixed = s_type == 'fixed-term contract employee'
                    mismatch = intern_yes & (~is_fixed)
                    n_mismatch = int(mismatch.sum())
                    if n_mismatch:
                        print(f"\nValidation: Found {n_mismatch} rows where internship='Yes' but type_of_employment != 'Fixed-Term Contract employee'")
                        # show identifying columns for flagged rows (use indices when available)
                        id_idx = headers.get('response_id', {}).get('index')
                        ts_idx = headers.get('timestamp', {}).get('index')
                        show_idxs = [i for i in [id_idx, ts_idx, col_intern_idx, col_type_idx] if i is not None]
                        flagged = df.loc[mismatch]
                        # select columns by integer positions
                        try:
                            print(flagged.iloc[:, show_idxs].to_string(index=False))
                        except Exception:
                            # fallback to showing the problematic columns by label if iloc selection fails
                            cols_by_label = [headers[k]['header'] for k in [key_intern, key_type] if k in headers]
                            id_col = headers.get('response_id', {}).get('header')
                            ts_col = headers.get('timestamp', {}).get('header')
                            show_cols = [c for c in [id_col, ts_col] + cols_by_label if c]
                            print(flagged.loc[:, show_cols].to_string(index=False))
                    else:
                        print('\nValidation: No internship/type_of_employment mismatches found.')
                else:
                    print('\nValidation: required columns not present in headers mapping or no data; skipping rule check.')
            except Exception as verr:
                print('Validation check failed:', verr)

            # optional: highlight flagged rows in a copy of the workbook
            highlight_out = getattr(args, 'highlight_output', None)
            try:
                if highlight_out and load_workbook is not None and n_mismatch:
                    # need header_row to map df indices to Excel rows; get via XML detection if missing
                    header_row_used = None
                    # read_headers_from_excel returns header_row as third element; try to get it
                    try:
                        # if earlier read used XML path, header_row variable may be in scope via outer returns
                        # re-run detect if necessary
                        if header_row:
                            header_row_used = header_row
                    except Exception:
                        header_row_used = None

                    if header_row_used is None:
                        try:
                            header_row_used, _ = detect_header_row_from_xml(path)
                        except Exception:
                            header_row_used = None

                    if header_row_used is None:
                        print('\nHighlighting skipped: cannot determine header row for mapping dataframe rows to Excel rows.')
                    else:
                        # flagged indices in dataframe 'mismatch' mask
                        flagged_idxs = list(df.index[mismatch])
                        try:
                            # Try to open original workbook and highlight in-place
                            wb = None
                            try:
                                wb = load_workbook(path)
                            except Exception:
                                wb = None

                            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                            if wb is not None and getattr(wb, 'worksheets', None) and len(wb.worksheets) > 0:
                                # choose sheet
                                if isinstance(sheet_val, int):
                                    try:
                                        ws = wb.worksheets[sheet_val]
                                    except Exception:
                                        ws = wb.active
                                else:
                                    if sheet_val in wb.sheetnames:
                                        ws = wb[sheet_val]
                                    else:
                                        ws = wb.active

                                maxcol = ws.max_column
                                for ridx in flagged_idxs:
                                    excel_row = header_row_used + 1 + int(ridx)
                                    for col in range(1, maxcol + 1):
                                        cell = ws.cell(row=excel_row, column=col)
                                        cell.fill = fill
                                wb.save(highlight_out)
                                print(f"\nHighlighted {len(flagged_idxs)} row(s) and wrote to: {highlight_out}")
                            else:
                                # fallback: create a new workbook from the dataframe and highlight flagged rows there
                                from openpyxl import Workbook
                                wb2 = Workbook()
                                ws2 = wb2.active
                                # write header row from df columns
                                headers_row = list(df.columns)
                                ws2.append(headers_row)
                                for r in df.itertuples(index=False, name=None):
                                    ws2.append(list(r))
                                maxcol2 = ws2.max_column
                                # in this new workbook, header is row 1
                                for ridx in flagged_idxs:
                                    excel_row = 1 + 1 + int(ridx)
                                    for col in range(1, maxcol2 + 1):
                                        ws2.cell(row=excel_row, column=col).fill = fill
                                wb2.save(highlight_out)
                                print(f"\nCreated highlighted workbook (new) with {len(flagged_idxs)} row(s): {highlight_out}")
                        except Exception as he:
                            print('Failed to write highlighted workbook:', he)
                elif highlight_out and load_workbook is None:
                    print('\nHighlighting requested but openpyxl is not available in the environment.')
            except Exception as hx:
                print('Highlight flow failed:', hx)
        else:
            print('No headers mapping produced.')
    except FileNotFoundError as fnf:
        print(fnf)
    except Exception as e:
        print('Error:', e)


if __name__ == '__main__':
    main()
